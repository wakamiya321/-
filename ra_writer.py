from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins
import pdfkit
from flask import render_template

# Flask外でも動くように、configは遅延評価
try:
    _WKHTMLTOPDF_CFG = pdfkit.configuration(wkhtmltopdf='/usr/bin/wkhtmltopdf')
except Exception:
    _WKHTMLTOPDF_CFG = None


HEAD_FONT = Font(name='ＭＳ Ｐゴシック', size=11, bold=True)
BODY_FONT = Font(name='ＭＳ Ｐゴシック', size=10)
THIN = Side(style='thin', color='000000')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _rows(ra: dict):
    # 表示順序
    order = [
        '作業場所','作業日','使用製品名','主な成分','SDSの有無',
        'GHSの分類（区分付き）','リスクレベルの判定','主なリスクの内容','リスク低減措置の検討',
        '作成者','所属会社'
    ]
    for k in order:
        v = ra.get(k, '')
        if isinstance(v, (list, tuple)):
            v = '\n'.join(v)
        yield k, v


def make_excel_bytes(ra: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = 'リスクアセスメント'

    # ページ設定（A4縦1枚に収める）
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 70

    ws['A1'] = '項目'; ws['B1'] = '内容'
    ws['A1'].font = HEAD_FONT; ws['B1'].font = HEAD_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['A1'].border = BORDER; ws['B1'].border = BORDER

    r = 2
    for k, v in _rows(ra):
        ws.cell(row=r, column=1, value=k).font = BODY_FONT
        c = ws.cell(row=r, column=2, value=v)
        c.font = BODY_FONT
        c.alignment = Alignment(wrapText=True, vertical='top')
        ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        # 目安の行高（内容の長さに応じて少し伸ばす）
        base = 18
        extra = min(120, int(len(str(v)) / 35) * 6)
        ws.row_dimensions[r].height = base + extra
        r += 1

    # バイナリへ
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), ra.get('filename_xlsx') or 'risk_assessment.xlsx'


def render_pdf_bytes(ra: dict):
    # HTMLテンプレートを利用してPDF化
    # ここではFlaskのJinja2テンプレートを直接呼ぶ想定。
    # app.py から import せず、render_template を事前に Flask context で呼ぶ形にしている。
    html = render_template('pdf.html', ra=ra)
    options = {
        'page-size': 'A4',
        'margin-top': '12mm',
        'margin-right': '10mm',
        'margin-bottom': '12mm',
        'margin-left': '10mm',
        'encoding': 'UTF-8',
        'enable-local-file-access': None,
        'print-media-type': None,
    }
    if _WKHTMLTOPDF_CFG is None:
        # wkhtmltopdf が無い環境ではPDF生成をスキップ（Excelのみ）
        return b"", ra.get('filename_pdf') or 'risk_assessment.pdf'

    pdf_bytes = pdfkit.from_string(html, False, options=options, configuration=_WKHTMLTOPDF_CFG)
    return pdf_bytes, ra.get('filename_pdf') or 'risk_assessment.pdf'
